"""Taxually OSS/VAT XLSX export — one row per invoice document.

Format: 20-column XLSX, sheet name "Your data".
"""
from __future__ import annotations

import datetime
import logging
import os
from decimal import Decimal
from pathlib import Path
from typing import Iterable

import openpyxl

from jtl2datev.core.models import RawInvoice
from jtl2datev.core.tax_engine import EU_COUNTRIES, looks_like_valid_vat_id, normalise_vat_id

logger = logging.getLogger(__name__)

TAXUALLY_COLUMNS: tuple[str, ...] = (
    "Transaction type",
    "Subject of the transaction",
    "Sales channel",
    "VAT number",
    "Transaction date",
    "Invoice number",
    "Departure country",
    "Customer's country",
    "Currency",
    "Gross amount",
    "VAT reporting country",
    "VAT Rate",
    "Net amount",
    "VAT amount",
    "Invoice date",
    "Local currency",
    "Exchange rate",
    "Gross amount_local",
    "Net amount_local",
    "VAT amount_local",
)

assert len(TAXUALLY_COLUMNS) == 20, f"Expected 20 columns, got {len(TAXUALLY_COLUMNS)}"

_DATE_FORMAT = "DD.MM.YYYY"


def _vat_reporting_country(
    customer_country: str,
    dispatch_country: str,
    vat_rate: float,
) -> str:
    if vat_rate > 0:
        return customer_country
    if customer_country == "GB":
        return "GB"
    return dispatch_country


def _invoice_date(d: datetime.date | datetime.datetime) -> datetime.date:
    if isinstance(d, datetime.datetime):
        return d.date()
    return d


def _aggregate_lines(invoice: RawInvoice) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """Return (total_gross, total_net, total_vat, first_vat_rate)."""
    if not invoice.lines:
        return Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0")
    total_gross = sum((abs(line.gross) for line in invoice.lines), Decimal("0"))
    total_net = sum((abs(line.net) for line in invoice.lines), Decimal("0"))
    total_vat = sum((abs(line.vat_amount) for line in invoice.lines), Decimal("0"))
    first_rate = invoice.lines[0].vat_rate
    return total_gross, total_net, total_vat, first_rate


def _build_taxually_workbook(invoices: Iterable[RawInvoice]) -> tuple[openpyxl.Workbook, int]:
    """Build a Taxually workbook from invoices. Returns (workbook, rows_written)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Your data"

    ws.append(list(TAXUALLY_COLUMNS))

    _date_col_indices = {
        TAXUALLY_COLUMNS.index("Transaction date") + 1,
        TAXUALLY_COLUMNS.index("Invoice date") + 1,
    }

    rows_written = 0

    for invoice in invoices:
        total_gross, _total_net, _total_vat, first_rate = _aggregate_lines(invoice)
        if invoice.is_credit_note:
            total_gross = -total_gross

        dispatch_country = invoice.warehouse_country.upper()
        customer_country = invoice.ship_to.country_iso.upper()
        currency = invoice.currency or "EUR"
        transaction_date = _invoice_date(invoice.invoice_date)
        transaction_type = "REFUND" if invoice.is_credit_note else "SALE"
        vat_rate_float = float(first_rate or Decimal("0"))
        vatc = _vat_reporting_country(customer_country, dispatch_country, vat_rate_float)
        vat_rate_normalised = vat_rate_float / 100.0

        # Customer-VAT-ID nur für IC-Supply-Fälle:
        # 0%-Steuersatz, Zielland EU oder XI (Nordirland), kein Export nach UK/CH.
        # Bei B2C ohne VAT-ID, bei UK/CH-Export, bei Standardsatz: leer.
        customer_vat_id: str | None = None
        if (
            vat_rate_float == 0
            and customer_country in (EU_COUNTRIES | {"XI"})
            and customer_country not in {"GB", "CH"}
            and invoice.ship_to.vat_id
        ):
            raw_vat = invoice.ship_to.vat_id
            if looks_like_valid_vat_id(raw_vat):
                customer_vat_id = normalise_vat_id(raw_vat, customer_country)
            else:
                logger.warning(
                    "invoice %s: ship_to.vat_id %r does not look like a valid VAT-ID — omitting",
                    invoice.invoice_no,
                    raw_vat,
                )

        row = [
            transaction_type,       # Transaction type
            "Goods",                # Subject of the transaction
            "Marketplace",          # Sales channel
            customer_vat_id,        # VAT number (Kunden-VAT bei IC-Supply)
            transaction_date,       # Transaction date
            invoice.invoice_no,     # Invoice number
            dispatch_country,       # Departure country
            customer_country,       # Customer's country
            currency,               # Currency
            float(total_gross),     # Gross amount
            vatc,                   # VAT reporting country
            vat_rate_normalised,    # VAT Rate
            None,                   # Net amount (Taxually calculates)
            None,                   # VAT amount (Taxually calculates)
            None,                   # Invoice date
            None,                   # Local currency
            None,                   # Exchange rate
            None,                   # Gross amount_local
            None,                   # Net amount_local
            None,                   # VAT amount_local
        ]

        assert len(row) == 20

        ws.append(row)
        rows_written += 1

        current_row = ws.max_row
        for col_idx in _date_col_indices:
            cell = ws.cell(row=current_row, column=col_idx)
            if cell.value is not None:
                cell.number_format = _DATE_FORMAT

    return wb, rows_written


def to_taxually_xlsx_bytes(invoices: Iterable[RawInvoice]) -> tuple[bytes, int]:
    """Build Taxually XLSX as bytes via BytesIO.

    Returns (payload_bytes, rows_written). Does not touch the filesystem.
    """
    import io as _io

    wb, rows_written = _build_taxually_workbook(invoices)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), rows_written


def format_taxually_xlsx(invoices: Iterable[RawInvoice], output_path: Path) -> int:
    """Write Taxually XLSX from invoices.

    Returns the number of data rows written.
    """
    payload, rows_written = to_taxually_xlsx_bytes(invoices)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = Path(str(output_path) + ".tmp")
    try:
        tmp.write_bytes(payload)
        os.replace(tmp, output_path)
    except Exception:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass
        raise
    logger.info("Taxually XLSX written: %d rows → %s", rows_written, output_path)
    return rows_written


def taxually_filename(year: int, month: int) -> str:
    """Return the conventional Taxually filename, e.g. Taxually-2026-01.xlsx."""
    return f"Taxually-{year}-{month:02d}.xlsx"
