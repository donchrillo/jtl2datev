"""Taxually XLSX export for Amazon intra-Community movements (Verbringungen).

Produces the same 20-column format as the regular Taxually export so the file
can be uploaded to Taxually alongside normal sales data.
"""
from __future__ import annotations

import logging
import os
from decimal import Decimal
from pathlib import Path

import openpyxl

from jtl2datev.core.config import OWN_VAT_IDS_VERBRINGUNG
from jtl2datev.core.taxually import TAXUALLY_COLUMNS
from jtl2datev.core.verbringung_parser import MovementRow
from jtl2datev.core.verbringung_pricing import PricingResult

logger = logging.getLogger(__name__)

_DATE_FORMAT = "DD.MM.YYYY"


def _transaction_date(row: MovementRow) -> str:
    """Return transaction date as DD.MM.YYYY string.

    For FC_TRANSFER: use depart_date (fall back to complete_date).
    For INBOUND: use complete_date (depart_date is the shipment start, complete_date
    is when the shipment actually arrived at the FC — that date appears in the
    reference XLSX).
    """
    if row.transaction_type == "INBOUND":
        d = row.complete_date or row.depart_date
    else:
        d = row.depart_date or row.complete_date
    if d is None:
        return ""
    return d.strftime("%d.%m.%Y")


def _gross_amount(row: MovementRow, pricing: dict[str, PricingResult]) -> float:
    pr = pricing.get(row.seller_sku)
    if pr is None or pr.ek_netto is None:
        return 0.0
    return float(Decimal(str(pr.ek_netto)) * row.qty)


def format_verbringung_xlsx(
    movements: list[MovementRow],
    pricing: dict[str, PricingResult],
    output_path: Path,
    own_vat_ids: dict[str, str] | None = None,
) -> int:
    """Write Taxually-format XLSX for intra-Community movements.

    Returns number of data rows written.
    """
    if own_vat_ids is None:
        own_vat_ids = OWN_VAT_IDS_VERBRINGUNG

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Your data"

    ws.append(list(TAXUALLY_COLUMNS))

    date_col_idx = TAXUALLY_COLUMNS.index("Transaction date") + 1

    missing_ek: list[str] = []
    rows_written = 0

    for row in movements:
        if row.transaction_type == "FC_TRANSFER":
            transaction_type = "Inventory transfer"
        else:
            transaction_type = "Sales"

        vat_number = own_vat_ids.get(row.departure_country, "")
        transaction_date = _transaction_date(row)
        invoice_number = row.transaction_event_id[:30]
        gross = _gross_amount(row, pricing)

        pr = pricing.get(row.seller_sku)
        if pr is None or pr.ek_netto is None:
            missing_ek.append(row.seller_sku)

        data_row = [
            transaction_type,           # Transaction type
            "Goods",                    # Subject of the transaction
            "",                         # Sales channel
            vat_number or None,         # VAT number
            transaction_date,           # Transaction date
            invoice_number,             # Invoice number
            row.departure_country,      # Departure country
            row.arrival_country,        # Customer's country
            "EUR",                      # Currency
            gross,                      # Gross amount
            None,                       # VAT reporting country
            None,                       # VAT Rate
            None,                       # Net amount
            None,                       # VAT amount
            None,                       # Invoice date
            None,                       # Local currency
            None,                       # Exchange rate
            None,                       # Gross amount_local
            None,                       # Net amount_local
            None,                       # VAT amount_local
        ]

        assert len(data_row) == 20

        ws.append(data_row)
        rows_written += 1

        if transaction_date:
            ws.cell(row=ws.max_row, column=date_col_idx).number_format = _DATE_FORMAT

    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = Path(str(output_path) + ".tmp")
    try:
        wb.save(str(tmp))
        os.replace(tmp, output_path)
    except Exception:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass
        raise

    if missing_ek:
        logger.warning(
            "format_verbringung_xlsx: %d rows written with gross=0.0 (no EK found): %s",
            len(missing_ek),
            missing_ek[:10],
        )
    logger.info(
        "Verbringung XLSX written: %d rows → %s (%d missing EK)",
        rows_written,
        output_path,
        len(missing_ek),
    )
    return rows_written
